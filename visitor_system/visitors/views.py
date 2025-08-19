# visitors/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.cache import never_cache, cache_page

from django.contrib import messages
from .models import Visit, Guest, StudentVisit, EmployeeProfile, Department, \
    STATUS_CHECKED_IN, STATUS_CHECKED_OUT, STATUS_AWAITING_ARRIVAL, STATUS_CANCELLED
from .forms import GuestRegistrationForm, StudentVisitRegistrationForm, HistoryFilterForm, ProfileSetupForm, \
    GuestInvitationFillForm, GuestInvitationFinalizeForm, GroupVisitRegistrationForm
from .models import GuestInvitation
from notifications.utils import send_guest_arrival_email # Импорт функции уведомления
from django.db.models import Q # Для сложных запросов
from django.contrib.auth.models import User
import datetime
from datetime import timedelta
import json
import logging
import uuid
from django.urls import reverse
from django.core import signing
from qrcode import QRCode
from io import BytesIO
from django.core.mail import send_mail
import os # Add this import
from django.conf import settings

from django.contrib.auth.models import Group # Импорт модели группы пользователей
from django.db.models import Count, Avg, F, DurationField
from django.db.models.functions import TruncDate, TruncMonth
from .filters import VisitFilter # Импорт фильтров для поиска
from itertools import chain
from operator import attrgetter
from collections import defaultdict
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.views.decorators.http import require_POST # Для ограничения методов
from rest_framework.views import APIView  # type: ignore
from rest_framework.response import Response  # type: ignore
from rest_framework.throttling import AnonRateThrottle, UserRateThrottle  # type: ignore
from django.utils.decorators import method_decorator
from django.utils.http import url_has_allowed_host_and_scheme # Для безопасного редиректа
from django.core.cache import cache
from django.views.decorators.vary import vary_on_cookie

from django.contrib.staticfiles.views import serve as serve_static

from notifications.utils import send_new_visit_notification_to_security

from .models import GroupInvitation, GroupGuest
from django.forms import modelformset_factory
from django.utils.crypto import get_random_string

logger = logging.getLogger(__name__)

# Имя группы для сотрудников ресепшн
RECEPTION_GROUP_NAME = "Reception"

FUNCTIONAL_ACCESS_GROUP_NAME = "FunctionalManager" # Например, "FunctionalManager" или "РасширенныйДоступ"

# --- Вспомогательная функция для получения визитов с учетом прав доступа ---
def get_scoped_visits_qs(user):
    """
    Возвращает QuerySet'ы для официальных и студенческих визитов,
    отфильтрованные в соответствии с правами пользователя.
    - Администраторы (is_staff) и члены группы RECEPTION_GROUP_NAME видят все визиты.
    - Остальные сотрудники видят визиты только своего департамента.
    """
    official_visits_qs = (
        Visit.objects.select_related('guest', 'employee', 'department', 'registered_by')
    )
    student_visits_qs = (
        StudentVisit.objects.select_related('guest', 'department', 'registered_by')
    )

    is_reception = user.is_authenticated and user.groups.filter(name=RECEPTION_GROUP_NAME).exists()
    is_staff = user.is_staff

    if not is_reception and not is_staff:
        if not user.is_authenticated or user.pk is None:
            official_visits_qs = official_visits_qs.none()
            student_visits_qs = student_visits_qs.none()
        else:
            try:
                # Пытаемся получить профиль сотрудника и его департамент
                employee_profile = EmployeeProfile.objects.get(user=user)
                if employee_profile.department:
                    employee_department = employee_profile.department
                    official_visits_qs = official_visits_qs.filter(department=employee_department)
                    student_visits_qs = student_visits_qs.filter(department=employee_department)
                else: # Сотрудник без департамента не видит визиты других департаментов
                    official_visits_qs = official_visits_qs.none()
                    student_visits_qs = student_visits_qs.none()
            except EmployeeProfile.DoesNotExist: # Пользователь без профиля сотрудника
                official_visits_qs = official_visits_qs.none()
                student_visits_qs = student_visits_qs.none()
    
    return official_visits_qs, student_visits_qs


def combine_visit_lists(official_visits_qs, student_visits_qs):
    """
    Объединяет два QuerySet'а (официальные и студенческие визиты) в один отсортированный список.
    Добавляет атрибут visit_kind для различения типов визитов в шаблоне.
    Сортирует по релевантному времени (entry_time или expected_entry_time).
    
    Args:
        official_visits_qs: QuerySet с объектами Visit
        student_visits_qs: QuerySet с объектами StudentVisit
    
    Returns:
        list: Отсортированный список визитов обоих типов
    """
    # Добавляем атрибут для различения в шаблоне
    for v in official_visits_qs: v.visit_kind = 'official'
    for v in student_visits_qs: v.visit_kind = 'student'
    
    def get_sort_key(visit):
        # Сортируем по релевантному времени: фактическое > планируемое
        # None ставим в самый конец (самое старое время)
        relevant_time = None
        if visit.entry_time:
            relevant_time = visit.entry_time
        elif hasattr(visit, 'expected_entry_time') and visit.expected_entry_time:
            relevant_time = visit.expected_entry_time

        # Даем очень старую дату для None, чтобы они были последними при reverse=True
        very_old_time = timezone.make_aware(datetime.datetime.min, timezone.get_default_timezone())
        return relevant_time if relevant_time else very_old_time

    # Объединяем и сортируем списки визитов
    combined_list = sorted(
        chain(official_visits_qs.iterator(), student_visits_qs.iterator()),
        key=get_sort_key,
        reverse=True  # Самые свежие (по фактическому или планируемому входу) - вверху
    )
    
    return combined_list
# -----------------------------------------------------------------------

@login_required
def register_guest_view(request):
    if request.method == 'POST':
        form = GuestRegistrationForm(request.POST, prefix="guest") # Используем префикс
        if form.is_valid():
            # Получаем выбор времени регистрации ИЗ ФОРМЫ (из радиокнопок)
            registration_type = request.POST.get('guest_reg_time', 'now') # 'now' или 'later'

            # Проверяем, если 'later', то должно быть время
            if registration_type == 'later' and not form.cleaned_data.get('expected_entry_time'):
                form.add_error('expected_entry_time', 'Укажите планируемое время для регистрации заранее.')
            else:
                try:
                    visit = form.save(request.user, registration_type=registration_type)
                    try:
                        send_new_visit_notification_to_security(visit, 'official')
                    except Exception as e:
                        logger.error(f"Ошибка при вызове send_new_visit_notification_to_security для гостя: {e}")
                    # --------------------------------------------
                    messages.success(request, f"Визит гостя {visit.guest.full_name} успешно зарегистрирован ({'сейчас' if registration_type == 'now' else 'заранее'})!")
                    return redirect('employee_dashboard')
                except Exception as e:
                    messages.error(request, f"Ошибка при регистрации визита: {e}")
        else:
            messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    else:
        form = GuestRegistrationForm(prefix="guest")

    context = {'guest_form': form} # Передаем форму под именем guest_form
    return render(request, 'visitors/register_guest.html', context)
@login_required
def qr_scan_view(request):
    return render(request, 'visitors/qr_scan.html')


@login_required
@require_POST
def qr_resolve_view(request):
    # Простейший rate limit для QR: до 10 запросов/5с на IP
    try:
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        from django.core.cache import cache
        k = f"qr_rl_{ip}"
        c = cache.get(k, 0)
        if c and int(c) >= 10:
            return JsonResponse({'error': 'too_many_requests'}, status=429)
        cache.incr(k) if cache.get(k) else cache.set(k, 1, timeout=5)
    except Exception:
        pass
    """Обработчик результата сканирования. Поддерживает:
    - подписанные токены формата signer.dumps({kind, id, action})
    - прямые ссылки нашего приложения
    """
    raw = request.POST.get('code', '').strip()
    if not raw:
        messages.error(request, 'Пустой код')
        return redirect('qr_scan')

    # Пробуем распарсить как подписанный токен
    try:
        data = signing.Signer().unsign_object(raw)
        kind = data.get('kind')
        pk = int(data.get('id'))
        action = data.get('action', 'checkin')
    except Exception:
        # Фоллбэк: парсим как URL нашего сайта вида /visitors/checkin/<kind>/<id>/
        try:
            from urllib.parse import urlparse
            p = urlparse(raw)
            path = p.path or raw
            parts = [x for x in path.split('/') if x]
            # ожидаем .../visitors/checkin/<kind>/<id>
            kind = parts[-2]
            pk = int(parts[-1])
            action = 'checkin'
        except Exception:
            messages.error(request, 'Некорректный код/ссылка')
            return redirect('qr_scan')

    # Выполняем действие
    if action == 'checkin':
        return check_in_visit(request, kind, pk)
    elif action == 'checkout':
        if kind == 'official':
            return mark_guest_exit_view(request, pk)
        elif kind == 'student':
            return mark_student_exit_view(request, pk)
    messages.error(request, 'Неизвестное действие')
    return redirect('qr_scan')


@login_required
def qr_code_view(request, kind: str, pk: int):
    """Отдаёт PNG QR код с подписанной полезной нагрузкой.
    kind: official|student
    action: по умолчанию checkin
    """
    payload = {'kind': kind, 'id': pk, 'action': 'checkin'}
    token = signing.Signer().sign_object(payload)
    url = request.build_absolute_uri(reverse('qr_resolve'))

    qr = QRCode(border=1)
    qr.add_data(token)  # Можно изменить на url с query, но токен компактнее
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    bio = BytesIO()
    img.save(bio, format='PNG')
    return HttpResponse(bio.getvalue(), content_type='image/png')

# ----------------------------------------------------------------------

"""@login_required
@cache_page(60 * 5)  # Кэшируем страницу на 5 минут
@vary_on_cookie     # Варьируем кэш в зависимости от пользователя (т.к. содержимое зависит от прав доступа)
def current_guests_view(request):
    Отображение списка гостей, которые сейчас находятся в здании.
    # Фильтруем визиты по статусу 'CHECKED_IN'
    #current_visits = Visit.objects.filter(
    #    status=STATUS_CHECKED_IN # <-- Изменено условие фильтрации
    #).select_related(
    #    'guest', 'employee', 'department', 'registered_by' # Оптимизация запроса
    #)
    # --- Добавим фильтрацию для StudentVisit ---
    #current_student_visits = StudentVisit.objects.filter(
    #    status=STATUS_CHECKED_IN # <-- То же условие
    #).select_related(
    #    'guest', 'department', 'registered_by'
    #)
    
    user = request.user
    # Получаем базовые QuerySet'ы с учетом прав доступа
    official_visits_qs, student_visits_qs = get_scoped_visits_qs(user)

    # Теперь фильтруем их по статусу "В здании"
    current_official_visits = official_visits_qs.filter(status=STATUS_CHECKED_IN)
    current_student_visits = student_visits_qs.filter(status=STATUS_CHECKED_IN)

    # Добавляем атрибут для различения в шаблоне (если нужно)
    #for v in current_visits: v.visit_kind = 'official'
    for v in current_official_visits: v.visit_kind = 'official'
    for v in current_student_visits: v.visit_kind = 'student'    # Объединяем и сортируем (например, по времени входа)
    # Создаем функцию-ключ для сортировки, которая безопасно обрабатывает None
    def safe_entry_time_key(visit):
        # Если entry_time равен None, возвращаем минимальную дату для сортировки
        if visit.entry_time is None:
            return timezone.make_aware(datetime.datetime.min, timezone.get_default_timezone())
        return visit.entry_time

    combined_list = sorted(
        chain(current_official_visits, current_student_visits),
        key=safe_entry_time_key,  # Используем безопасную функцию-ключ
        reverse=True  # Самые недавние вверху
    )
    # ------------------------------------------

    context = {
        # Передаем объединенный список в шаблон
        'current_visits': combined_list
    }
    return render(request, 'visitors/current_guests.html', context)"""
# ---------------------------------

from django.views.decorators.cache import never_cache

@login_required
@never_cache  # Отключаем кэширование для динамического представления с фильтрацией
def visit_history_view(request):
    """Отображение И общей истории визитов с фильтрацией."""
    # Если есть GET-параметры, это фильтрация - не кэшируем
    user = request.user # Получаем текущего пользователя
    # Получаем визиты с учетом прав доступа пользователя
    official_visits_qs, student_visits_qs = get_scoped_visits_qs(user)    # Получаем групповые визиты - для всех, у кого есть права на историю визитов
    # Выбираем только зарегистрированные групповые визиты
    group_visits_qs = GroupInvitation.objects.filter(
        is_registered=True,  # Берем только зарегистрированные групповые визиты
        visit_time__isnull=False  # Убедимся, что у групповых визитов указано время
    ).select_related(
        'department', 'employee'
    ).prefetch_related('guests')
    
    # Фильтруем групповые визиты по правам доступа, аналогично другим типам визитов
    is_reception = user.is_authenticated and user.groups.filter(name=RECEPTION_GROUP_NAME).exists()
    is_staff = user.is_staff
    
    if not is_reception and not is_staff:
        if not user.is_authenticated or user.pk is None:
            group_visits_qs = group_visits_qs.none()
        else:
            try:
                employee_profile = EmployeeProfile.objects.get(user=user)
                if employee_profile.department:
                    # Если это обычный сотрудник с департаментом, показываем групповые визиты этого департамента
                    group_visits_qs = group_visits_qs.filter(department=employee_profile.department)
                else:
                    # Если нет департамента - не показываем групповые визиты других департаментов
                    group_visits_qs = group_visits_qs.none()
            except EmployeeProfile.DoesNotExist:
                group_visits_qs = group_visits_qs.none()
    
    # Определяем, должен ли пользователь видеть полные фильтры
    is_reception_or_staff = user.is_staff or user.groups.filter(name=RECEPTION_GROUP_NAME).exists()

    # Инициализируем форму фильтра GET-данными
    filter_form = HistoryFilterForm(request.GET or None)
    
    # Применяем фильтры, если форма была отправлена (есть GET параметры)
    if request.GET and filter_form.is_valid():
        logger.debug(f"Applying filters: {filter_form.cleaned_data}")
        # Общие фильтры
        guest_name_query = filter_form.cleaned_data.get('guest_name') # Corrected: was 'guests', changed variable name
        guest_iin = filter_form.cleaned_data.get('guest_iin')
        entry_date_from = filter_form.cleaned_data.get('entry_date_from')
        entry_date_to = filter_form.cleaned_data.get('entry_date_to')
        # Фильтрация по типу визита
        visit_type = filter_form.cleaned_data.get('visit_type')
        if visit_type:
            logger.debug(f"Filtering by visit_type: '{visit_type}'")
            if visit_type == 'official':
                # Если выбран только официальный тип, отключаем другие типы визитов
                student_visits_qs = student_visits_qs.none()
                group_visits_qs = group_visits_qs.none()
            elif visit_type == 'student':
                # Если выбран только студенческий тип, отключаем другие типы визитов
                official_visits_qs = official_visits_qs.none()
                group_visits_qs = group_visits_qs.none()
            elif visit_type == 'group':
                # Если выбран только групповой тип, отключаем другие типы визитов
                official_visits_qs = official_visits_qs.none()
                student_visits_qs = student_visits_qs.none()
                # Применяем фильтр по статусу
        status = filter_form.cleaned_data.get('status')
        if status:
            logger.debug(f"Filtering by status: '{status}'")
            official_visits_qs = official_visits_qs.filter(status=status)
            student_visits_qs = student_visits_qs.filter(status=status)
            # Для групповых визитов используем логику сопоставления статусов:
            if status == STATUS_CHECKED_IN:
                group_visits_qs = group_visits_qs.filter(is_registered=True, is_completed=False, visit_time__lte=timezone.now())
            elif status == STATUS_CHECKED_OUT:
                group_visits_qs = group_visits_qs.filter(is_registered=True, is_completed=True)
            elif status == STATUS_AWAITING_ARRIVAL:
                group_visits_qs = group_visits_qs.filter(is_registered=True, is_completed=False, visit_time__gt=timezone.now())
            elif status == STATUS_CANCELLED: # GroupInvitation doesn't have a CANCELLED status
                group_visits_qs = group_visits_qs.none()
            else: # For any other status from choices not explicitly handled
                group_visits_qs = group_visits_qs.none()
                
        department = filter_form.cleaned_data.get('department')

        # Apply guest_name_query filter
        if guest_name_query:
            official_visits_qs = official_visits_qs.filter(guest__full_name__icontains=guest_name_query)
            student_visits_qs = student_visits_qs.filter(guest__full_name__icontains=guest_name_query)
            # Filter group visits by guest name (any guest in the group)
            group_visits_qs = group_visits_qs.filter(guests__full_name__icontains=guest_name_query)
        
        if guest_iin:
            # Хэшируем ввод и ищем совпадения по полному ИИН через хэш и маске по последним 4
            import hashlib
            iin_hash = hashlib.sha256(guest_iin.encode()).hexdigest()
            official_visits_qs = official_visits_qs.filter(guest__iin_hash=iin_hash)
            student_visits_qs = student_visits_qs.filter(guest__iin_hash=iin_hash)
            group_visits_qs = group_visits_qs.filter(guests__iin__icontains=guest_iin)
        
        if entry_date_from:
            official_visits_qs = official_visits_qs.filter(entry_time__date__gte=entry_date_from)
            student_visits_qs = student_visits_qs.filter(entry_time__date__gte=entry_date_from)
            # Для групповых визитов используем visit_time
            group_visits_qs = group_visits_qs.filter(visit_time__date__gte=entry_date_from)
        
        if entry_date_to:
            official_visits_qs = official_visits_qs.filter(entry_time__date__lte=entry_date_to)
            student_visits_qs = student_visits_qs.filter(entry_time__date__lte=entry_date_to)
            # Для групповых визитов используем visit_time
            group_visits_qs = group_visits_qs.filter(visit_time__date__lte=entry_date_to)
        
        if department:
            official_visits_qs = official_visits_qs.filter(department=department)
            student_visits_qs = student_visits_qs.filter(department=department)
            group_visits_qs = group_visits_qs.filter(department=department)
        
        # Фильтры только для Visit
        employee_info = filter_form.cleaned_data.get('employee_info')
        if employee_info:
            official_visits_qs = official_visits_qs.filter(
                Q(employee__first_name__icontains=employee_info) |
                Q(employee__last_name__icontains=employee_info) |
                Q(employee__email__icontains=employee_info)
            )
            # Также применяем фильтр к групповым визитам
            group_visits_qs = group_visits_qs.filter(
                Q(employee__first_name__icontains=employee_info) |
                Q(employee__last_name__icontains=employee_info) |
                Q(employee__email__icontains=employee_info)
            )

        # Фильтры только для StudentVisit
        student_id = filter_form.cleaned_data.get('student_id_number')
        student_group = filter_form.cleaned_data.get('student_group')
        if student_id:
            student_visits_qs = student_visits_qs.filter(student_id_number__icontains=student_id)
        if student_group:
            student_visits_qs = student_visits_qs.filter(student_group__icontains=student_group)    # Добавляем атрибут для различения в шаблоне
    for v in official_visits_qs: v.visit_kind = 'official'
    for v in student_visits_qs: v.visit_kind = 'student'    # Добавляем аналогичный атрибут для групповых визитов
    for v in group_visits_qs: 
        v.visit_kind = 'group'
        # Добавляем требуемые атрибуты для совместимости с шаблоном
        if v.is_completed:
            v.status = STATUS_CHECKED_OUT
        else: # is_registered = True, visit_time is not null, and not is_completed
            if v.visit_time and v.visit_time > timezone.now():
                v.status = STATUS_AWAITING_ARRIVAL
            else: # visit_time is now or in the past, and not completed
                v.status = STATUS_CHECKED_IN
        v.entry_time = v.visit_time
        v.exit_time = v.exit_time  # Fixed: Use the model's exit_time field
        # Добавляем гостей как атрибуты для более удобного доступа
        guests_count = v.guests.count()
        if guests_count > 0:
            # Получаем первого гостя в группе для отображения в списке
            first_guest = v.guests.first()
            v.guest = first_guest
            v.guests_count = guests_count
        else:
            v.guest = None
            v.guests_count = 0
            
        # Добавляем ссылку на само групповое приглашение для доступа в шаблоне
        v.group_invitation = v
    
    def get_sort_key(visit):
        # Сортируем по релевантному времени: фактическое > планируемое
        # None ставим в самый конец (самое старое время)
        relevant_time = None
        if hasattr(visit, 'entry_time') and visit.entry_time:
            relevant_time = visit.entry_time
        elif hasattr(visit, 'expected_entry_time') and visit.expected_entry_time:
            relevant_time = visit.expected_entry_time
        elif hasattr(visit, 'visit_time') and visit.visit_time:
            relevant_time = visit.visit_time

        # Даем очень старую дату для None, чтобы они были последними при reverse=True
        very_old_time = timezone.make_aware(datetime.datetime.min, timezone.get_default_timezone())
        return relevant_time if relevant_time else very_old_time

    # Объединяем все три типа визитов в один список
    combined_list = sorted(
        chain(official_visits_qs, student_visits_qs, group_visits_qs),
        key=get_sort_key, # Используем функцию ключа
        reverse=True # Самые свежие - вверху
    )# ---------------------------------------------


    # --- Логика пагинации ---
    try:
        items_per_page = int(request.GET.get('per_page', 20))  # Получаем количество записей из GET параметра
        if items_per_page not in [10, 20, 50, 100]:  # Разрешаем только определенные значения
            items_per_page = 20  # Значение по умолчанию
    except (ValueError, TypeError):
        items_per_page = 20  # Если параметр невалидный, используем значение по умолчанию
        
    paginator = Paginator(combined_list, items_per_page)
    page_number = request.GET.get('page')  # Получаем номер страницы из GET параметра 'page'

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        # Если номер страницы не целое число, показать первую страницу
        page_obj = paginator.get_page(1)
    except EmptyPage:
        # Если номер страницы больше максимального, показать последнюю страницу
        page_obj = paginator.get_page(paginator.num_pages)
    # ------------------------

    # --- Формируем строку GET параметров для сохранения фильтров в пагинации ---
    # Удаляем параметр 'page', чтобы он не дублировался
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']
    filter_params_url = get_params.urlencode()
    # --------------------------------------------------------------------

    context = {
        #'all_visits': combined_list, # Передаем объединенный список
        'page_obj': page_obj,
        'filter_form': filter_form,   # Передаем форму для отображения
        'filter_params_url': filter_params_url,
        'show_filters': is_reception_or_staff # <--- Передаем флаг в шаблон
    }
    return render(request, 'visitors/visit_history.html', context)
# ---------------------------------

@login_required
def visit_detail_view(request, visit_id):
    """Отображение подробной информации о конкретном визите."""
    # Получаем объект Visit по ID или возвращаем 404, если не найден
    # Используем select_related для оптимизации запросов к связанным моделям
    visit = get_object_or_404(
        Visit.objects.select_related('guest', 'employee', 'department', 'registered_by'),
        pk=visit_id
    )

    # Можно добавить проверку прав доступа, если нужно
    # Например, чтобы только админ или участники визита могли его смотреть

    context = {
        'visit': visit
    }
    return render(request, 'visitors/visit_detail.html', context)
# ---------------------------------

# --- Представление для деталей визита студента ---
@login_required
def student_visit_detail_view(request, visit_id):
    """Отображение подробной информации о визите студента/абитуриента."""
    student_visit = get_object_or_404(
        StudentVisit.objects.select_related('guest', 'department', 'registered_by'),
        pk=visit_id
    )
    # TODO: Проверка прав доступа, если нужна

    context = {
        'student_visit': student_visit # Передаем объект в контекст
    }
    # Используем новый шаблон
    return render(request, 'visitors/student_visit_detail.html', context)
# ------------------------------------------------------

@login_required
@require_POST # Ограничиваем метод только POST-запросами
def mark_guest_exit_view(request, visit_id):
    """Отмечает выход гостя (модель Visit)."""
    
    # Пытаемся получить объект Visit по ID
    visit = get_object_or_404(Visit, pk=visit_id)

    # Проверяем, находится ли гость в здании перед отметкой выхода
    if visit.status != STATUS_CHECKED_IN:
        messages.warning(request, f"Невозможно отметить выход для '{visit.guest.full_name}'. "
                                    f"Текущий статус: {visit.get_status_display()}. Возможно, выход уже был отмечен ранее.")
        # Возвращаем пользователя туда, откуда он пришел (или на страницу текущих гостей)
        return redirect(request.META.get('HTTP_REFERER', 'visit_history'))        # Если все в порядке, отмечаем время выхода и меняем статус
    visit.exit_time = timezone.now()
    visit.status = STATUS_CHECKED_OUT
    visit.save(update_fields=['exit_time', 'status']) # Обновляем только нужные поля

    # Асинхронно отправляем уведомление о выходе, если у визита есть назначенный сотрудник с email
    if hasattr(visit, 'host_employee') and visit.host_employee and visit.host_employee.email:
        from notifications.tasks import send_exit_notification
        subject = f"Посетитель {visit.guest.full_name} покинул здание"
        message = f"Посетитель {visit.guest.full_name} покинул здание в {visit.exit_time.strftime('%H:%M:%S')}."
        
        # Вызываем задачу Celery асинхронно
        send_exit_notification.delay(
            visit.host_employee.email,
            subject,
            message
        )
        logger.info(f"Поставлена задача на отправку уведомления о выходе для визита {visit_id}")

    messages.success(request, f"Выход гостя '{visit.guest.full_name}' успешно зарегистрирован.")
    # Перенаправляем на список текущих гостей
    return redirect('visit_history')

    """except Http404:
        # Явно обрабатываем случай, когда визит не найден (ошибка 404)
        logger.warning(f"Attempted to mark exit for non-existent Visit ID: {visit_id} by user {request.user.username}")
        messages.error(request, f"Ошибка: Визит гостя с ID {visit_id} не найден. Возможно, он был удален или уже обработан.")
        # Перенаправляем на безопасную страницу, например, историю
        return redirect('visit_history')

    except Exception as e:
        # Ловим другие возможные ошибки
        logger.error(f"Unexpected error marking guest exit for visit ID {visit_id}: {e}", exc_info=True)
        messages.error(request, "Произошла непредвиденная ошибка при отметке выхода гостя.")
        return redirect(request.META.get('HTTP_REFERER', 'current_guests'))"""

# --- View для отметки выхода СТУДЕНТА ---
@login_required
@require_POST
def mark_student_exit_view(request, visit_id):
    """Отмечает выход студента/абитуриента (модель StudentVisit)."""
    try:
        # Пытаемся получить объект StudentVisit по ID
        visit = get_object_or_404(StudentVisit, pk=visit_id)

        # Проверяем статус
        if visit.status != STATUS_CHECKED_IN:
            messages.warning(request, f"Невозможно отметить выход для '{visit.guest.full_name}'. "
                                      f"Текущий статус: {visit.get_status_display()}. Возможно, выход уже был отмечен ранее.")
            return redirect(request.META.get('HTTP_REFERER', 'visit_history'))

        # Отмечаем выход
        visit.exit_time = timezone.now()
        visit.status = STATUS_CHECKED_OUT
        visit.save(update_fields=['exit_time', 'status'])

        messages.success(request, f"Выход посетителя '{visit.guest.full_name}' успешно зарегистрирован.")
        return redirect('visit_history')

    except Http404:
        logger.warning(f"Attempted to mark exit for non-existent StudentVisit ID: {visit_id} by user {request.user.username}")
        messages.error(request, f"Ошибка: Визит студента/абитуриента с ID {visit_id} не найден. Возможно, он был удален или уже обработан.")
        return redirect('visit_history')

    except Exception as e:
        logger.error(f"Unexpected error marking student exit for visit ID {visit_id}: {e}", exc_info=True)
        messages.error(request, "Произошла непредвиденная ошибка при отметке выхода посетителя.")
        return redirect(request.META.get('HTTP_REFERER', 'current_guests'))


@login_required
def employee_dashboard_view(request):
    """Панель управления для обычного сотрудника."""
    user = request.user
    
    today = timezone.now().date()
    one_week_later = today + timedelta(days=7)
    
    # Получаем департамент текущего сотрудника
    try:
        user_department = user.employee_profile.department
    except (AttributeError, EmployeeProfile.DoesNotExist):
        user_department = None
    # Получаем визиты на ближайшую неделю для сотрудников своего департамента
    if user_department:
        # Получаем всех сотрудников департамента
        department_employee_ids = User.objects.filter(
            employee_profile__department=user_department
        ).values_list('id', flat=True)
        
        # Получаем визиты всех сотрудников департамента
        upcoming_visits_qs = Visit.objects.filter(
            employee_id__in=department_employee_ids,  # Фильтруем по списку ID сотрудников департамента
            expected_entry_time__date__gte=today,
            expected_entry_time__date__lt=one_week_later,
            status__in=[STATUS_AWAITING_ARRIVAL, STATUS_CHECKED_IN],  # Ожидаемая регистрация и уже в здании
        ).select_related('guest', 'department', 'employee').order_by('expected_entry_time')
    else:
        # Если у пользователя нет департамента, показываем только его визиты
        upcoming_visits_qs = Visit.objects.filter(
            employee=user,
            expected_entry_time__date__gte=today,
            expected_entry_time__date__lt=one_week_later,
            status__in=[STATUS_AWAITING_ARRIVAL, STATUS_CHECKED_IN],  # Ожидаемая регистрация и уже в здании
        ).select_related('guest', 'department', 'employee').order_by('expected_entry_time')
    
    # Convert queryset to list and add visit_kind for template URL generation
    upcoming_visits_week_list = list(upcoming_visits_qs)
    for visit_obj in upcoming_visits_week_list:
        visit_obj.visit_kind = 'official'
    # Получаем студенческие визиты на ту же неделю
    if user_department:
        student_visits_qs = StudentVisit.objects.filter(
            department=user_department,
            entry_time__date__gte=today,
            entry_time__date__lt=one_week_later,
            status__in=[STATUS_AWAITING_ARRIVAL, STATUS_CHECKED_IN],
        ).select_related('guest', 'department').order_by('entry_time')
    else:
        student_visits_qs = StudentVisit.objects.filter(
            registered_by=user,
            entry_time__date__gte=today,
            entry_time__date__lt=one_week_later,
            status__in=[STATUS_AWAITING_ARRIVAL, STATUS_CHECKED_IN],
        ).select_related('guest', 'department').order_by('entry_time')
    
    # Добавляем студенческие визиты в список
    student_visits_list = list(student_visits_qs)
    for visit_obj in student_visits_list:
        visit_obj.visit_kind = 'student'
    # Получаем зарегистрированные групповые визиты на ту же неделю
    if user_department:
        group_visits_qs = GroupInvitation.objects.filter(
            department=user_department,
            visit_time__date__gte=today,
            visit_time__date__lt=one_week_later,
            is_registered=True,
            is_completed=False,  # Только незавершенные (не вышедшие)
        ).select_related('department', 'employee').prefetch_related('guests').order_by('visit_time')
    else:
        group_visits_qs = GroupInvitation.objects.filter(
            employee=user,
            visit_time__date__gte=today,
            visit_time__date__lt=one_week_later,
            is_registered=True,
            is_completed=False,  # Только незавершенные (не вышедшие)
        ).select_related('department', 'employee').prefetch_related('guests').order_by('visit_time')
    # Добавляем групповые визиты в список
    group_visits_list = list(group_visits_qs)
    for visit_obj in group_visits_list:
        visit_obj.visit_kind = 'group'
        # Добавляем совместимость с шаблоном
        visit_obj.expected_entry_time = visit_obj.visit_time
        visit_obj.status = STATUS_AWAITING_ARRIVAL  # Групповые визиты в ожидании прибытия
        # Добавляем ссылку на само групповое приглашение для шаблона
        visit_obj.group_invitation = visit_obj
        # Добавляем информацию о гостях для отображения
        guests_count = visit_obj.guests.count()
        if guests_count > 0:
            first_guest = visit_obj.guests.first()
            visit_obj.guest = first_guest
            visit_obj.guests_count = guests_count
        else:
            visit_obj.guest = None
            visit_obj.guests_count = 0

    # Объединяем все три списка и сортируем по времени входа
    upcoming_visits_week_list.extend(student_visits_list)
    upcoming_visits_week_list.extend(group_visits_list)
    upcoming_visits_week_list.sort(key=lambda x: getattr(x, 'expected_entry_time', None) or getattr(x, 'entry_time', None) or getattr(x, 'visit_time', None))

    # Получаем заполненные, но не зарегистрированные приглашения для текущего пользователя
    pending_invitations = GuestInvitation.objects.filter(
        employee=user,
        is_filled=True,
        is_registered=False
    ).order_by('-created_at')
    
    # Гости, которых ожидает данный сотрудник и которые еще не вышли
    my_current_guests = Visit.objects.filter(
        employee=user,
        exit_time__isnull=True
    ).select_related('guest', 'department')
    
    # История визитов, связанных с этим сотрудником (как принимающий или регистрирующий)
    my_visit_history = Visit.objects.filter(
        Q(employee=user) | Q(registered_by=user)
    ).select_related('guest', 'department', 'employee', 'registered_by').order_by('-entry_time')[:20] # Последние 20# Данные для графика активности визитов
    # Генерируем данные для разных периодов (сегодня, неделя, месяц)
    visit_chart_data = {}
    
    # Фильтр по текущему департаменту пользователя
    department_filter = Q(department=user_department) if user_department else Q(employee=user)
    # Для StudentVisit нет employee, поэтому создаем отдельный фильтр
    student_department_filter = Q(department=user_department) if user_department else Q()
    
    # Данные для графика по часам (сегодня)
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    hourly_data = [0] * 24  # Инициализируем нулями для каждого часа
    
    # Получаем данные для официальных визитов по часам
    today_visits = Visit.objects.filter(
        department_filter, 
        entry_time__gte=today_start,
        entry_time__lt=today_end
    ).values('entry_time__hour').annotate(count=Count('id'))
    
    # Добавляем данные студенческих визитов
    today_student_visits = StudentVisit.objects.filter(
        student_department_filter,
        entry_time__gte=today_start,
        entry_time__lt=today_end
    ).values('entry_time__hour').annotate(count=Count('id'))
    
    # Суммируем данные официальных визитов
    for visit in today_visits:
        hour = visit['entry_time__hour']
        if 0 <= hour < 24:  # Проверка на всякий случай
            hourly_data[hour] += visit['count']
    
    # Суммируем данные студенческих визитов
    for visit in today_student_visits:
        hour = visit['entry_time__hour']
        if 0 <= hour < 24:  # Проверка на всякий случай
            hourly_data[hour] += visit['count']
    
    visit_chart_data['today'] = hourly_data
    # Данные для графика по дням недели
    week_start = today_start - timedelta(days=today_start.weekday())  # Понедельник текущей недели
    week_end = week_start + timedelta(days=7)
    weekly_data = [0] * 7  # Пн, Вт, Ср, Чт, Пт, Сб, Вс
    
    # Получаем данные официальных визитов по дням недели
    week_visits = Visit.objects.filter(
        department_filter,
        entry_time__gte=week_start,
        entry_time__lt=week_end
    ).values('entry_time__week_day').annotate(count=Count('id'))
    
    # Получаем данные студенческих визитов по дням недели
    week_student_visits = StudentVisit.objects.filter(
        student_department_filter,
        entry_time__gte=week_start,
        entry_time__lt=week_end
    ).values('entry_time__week_day').annotate(count=Count('id'))
    
    # В Django week_day: 1=воскресенье, 2=понедельник, ..., 7=суббота
    # Преобразуем в нашу индексацию: 0=понедельник, ..., 6=воскресенье
    for visit in week_visits:
        day = visit['entry_time__week_day']
        # Переводим из Django недельного дня в индекс массива (0-6, Пн-Вс)
        idx = (day - 2) % 7
        weekly_data[idx] += visit['count']
    
    # Добавляем данные студенческих визитов
    for visit in week_student_visits:
        day = visit['entry_time__week_day']
        # Переводим из Django недельного дня в индекс массива (0-6, Пн-Вс)
        idx = (day - 2) % 7
        weekly_data[idx] += visit['count']
    
    visit_chart_data['week'] = weekly_data
    # Данные для графика по дням месяца
    month_start = today_start.replace(day=1)  # Первый день текущего месяца
    next_month = month_start.month + 1 if month_start.month < 12 else 1
    next_month_year = month_start.year if month_start.month < 12 else month_start.year + 1
    month_end = month_start.replace(year=next_month_year, month=next_month, day=1)
    
    # Инициализируем массив с нулями для каждого дня месяца (максимум 31 день)
    monthly_data = [0] * 31
    
    # Получаем данные официальных визитов по дням месяца
    month_visits = Visit.objects.filter(
        department_filter,
        entry_time__gte=month_start,
        entry_time__lt=month_end
    ).values('entry_time__day').annotate(count=Count('id'))
    
    # Получаем данные студенческих визитов по дням месяца
    month_student_visits = StudentVisit.objects.filter(
        student_department_filter,
        entry_time__gte=month_start,
        entry_time__lt=month_end
    ).values('entry_time__day').annotate(count=Count('id'))
    
    # Суммируем данные официальных визитов
    for visit in month_visits:
        day = visit['entry_time__day']
        if 1 <= day <= 31:  # Проверка на всякий случай
            monthly_data[day-1] += visit['count']  # -1 т.к. индексы начинаются с 0
    
    # Суммируем данные студенческих визитов
    for visit in month_student_visits:
        day = visit['entry_time__day']
        if 1 <= day <= 31:  # Проверка на всякий случай
            monthly_data[day-1] += visit['count']  # -1 т.к. индексы начинаются с 0
    
    visit_chart_data['month'] = monthly_data# Конвертируем в JSON для передачи в шаблон
    import json
    visit_chart_data_json = json.dumps(visit_chart_data)
    # Получаем недавние визиты (включая уже покинувших посетителей) 
    # за последние 7 дней из текущего департамента
    recent_time_threshold = timezone.now() - timedelta(days=7)
    
    if user_department:
        # Если есть департамент, фильтруем по нему
        recent_official_visits = Visit.objects.filter(
            department=user_department,
            entry_time__gte=recent_time_threshold
        ).select_related('guest', 'department', 'employee').order_by('-entry_time')[:10]
        
        recent_student_visits = StudentVisit.objects.filter(
            department=user_department,
            entry_time__gte=recent_time_threshold
        ).select_related('guest', 'department').order_by('-entry_time')[:10]
    else:
        # Если нет департамента, показываем визиты, связанные с пользователем
        recent_official_visits = Visit.objects.filter(
            Q(employee=user) | Q(registered_by=user),
            entry_time__gte=recent_time_threshold
        ).select_related('guest', 'department', 'employee').order_by('-entry_time')[:10]
        
        recent_student_visits = StudentVisit.objects.filter(
            registered_by=user,
            entry_time__gte=recent_time_threshold
        ).select_related('guest', 'department').order_by('-entry_time')[:10]
    
    # Добавляем атрибут для различения типов визитов в шаблоне
    for v in recent_official_visits: v.visit_kind = 'official'
    for v in recent_student_visits: v.visit_kind = 'student'    # Объединяем и сортируем по времени входа (самые новые вверху)
    recent_visits = sorted(
        chain(recent_official_visits, recent_student_visits),
        key=attrgetter('entry_time'),
        reverse=True
    )[:10]  # Берем только 10 самых недавних визитов
    
    # Инициализируем переменные до try-except блока
    has_invitation_data_issues = False
    group_invitations = []
    
    # Рассчитаем процентное соотношение типов визитов для графика
    official_count = Visit.objects.filter(department_filter).count()
    student_count = StudentVisit.objects.filter(student_department_filter).count()
    total_count = official_count + student_count
    
    official_percent = int((official_count / total_count) * 100) if total_count > 0 else 50
    student_percent = int((student_count / total_count) * 100) if total_count > 0 else 50
    try:
        # Получаем групповые приглашения с предварительной загрузкой данных о гостях
        # Изменяем фильтрацию для получения и заполненных, и незаполненных приглашений
        group_invitations = GroupInvitation.objects.filter(
            employee=user,
            is_filled=True,  # Получаем только заполненные
            is_registered=False  # Которые еще не зарегистрированы
        ).select_related('department').prefetch_related('guests').order_by('-created_at')
        
        # Для каждого приглашения безопасно вычисляем количество гостей и добавляем как атрибут
        for invite in group_invitations:
            try:
                guest_count = invite.guests.count()
                invite.guest_count = guest_count
                
                # Проверяем обязательные поля
                if not invite.token:  # Удален чек invite.department
                    logger.warning(f"Приглашение {invite.pk} имеет некорректные данные (отсутствует token)")
                    has_invitation_data_issues = True
            except Exception as inner_e:
                logger.warning(f"Не удалось получить количество гостей для приглашения {invite.pk}: {inner_e}")
                invite.guest_count = 0
                has_invitation_data_issues = True
            
    except Exception as e:
        logger.error(f"Ошибка при получении групповых приглашений: {e}", exc_info=True)
        group_invitations = []
        has_invitation_data_issues = True
    
    # Если были проблемы с данными, добавим предупреждение для пользователя
    if has_invitation_data_issues:
        messages.warning(request, "Некоторые групповые приглашения могут отображаться некорректно из-за проблем с данными.")
    
    context = {
        'upcoming_visits': upcoming_visits_week_list, # Визиты на неделю всего департамента
        'my_current_guests': my_current_guests,
        'my_visit_history': my_visit_history,
        'recent_visits': recent_visits,  # Добавляем недавние визиты в контекст
        'today': today,  # Добавляем переменную today для условного форматирования в шаблоне
        'department_name': user_department.name if user_department else "Нет департамента",
        'visit_chart_data': visit_chart_data_json,  # Добавляем данные для графика
        'pending_invitations': pending_invitations,  # Добавляем заполненные приглашения
        'group_invitations': group_invitations,  # Добавляем приглашения в группы
        'has_invitation_data_issues': has_invitation_data_issues,  # Флаг проблем с данными приглашений
        'official_visits_percent': official_percent,  # Процент официальных визитов для графика
        'student_visits_percent': student_percent,   # Процент студенческих визитов для графика
    }
    return render(request, 'visitors/employee_dashboard.html', context)

def has_functional_access(user):
    """
    Проверяет, имеет ли пользователь расширенный функциональный доступ.
    Доступ предоставляется, если:
    1. Пользователь аутентифицирован.
    2. Пользователь является администратором (is_staff = True).
    3. Пользователь состоит в группе FUNCTIONAL_ACCESS_GROUP_NAME.
    """
    if not user.is_authenticated:
        return False
    if user.is_staff: # Администраторы всегда имеют доступ
        return True
    try:
        return user.groups.filter(name=FUNCTIONAL_ACCESS_GROUP_NAME).exists()
    except Group.DoesNotExist: # На случай, если группа была удалена или еще не создана
        logger.warning(f"Группа '{FUNCTIONAL_ACCESS_GROUP_NAME}' не найдена в базе данных. "
                       f"Функциональный доступ для пользователя {user.username} не предоставлен по группе.")
        return False
# -----------------------------------------

# --- Пример использования нового уровня доступа ---
@login_required
@user_passes_test(has_functional_access) # Используем новую проверку
def example_special_feature_view(request):
    """
    Пример представления, доступного администраторам (is_staff)
    и пользователям из группы FUNCTIONAL_ACCESS_GROUP_NAME.
    """
    # Ваша логика для этой специальной функции
    context = {
        'message': f"Добро пожаловать, {request.user.username}! У вас есть доступ к этой специальной функции."
    }
    return render(request, 'visitors/special_feature_example.html', context) # Создайте этот шаблон
# --------------------------------------------------

# Проверка, является ли пользователь администратором (staff)
def is_admin(user):
    return user.is_authenticated and user.is_staff

@user_passes_test(is_admin) # Доступ только для администраторов
def admin_dashboard_view(request):
    """Панель управления для администратора."""
    current_visits = Visit.objects.filter(exit_time__isnull=True).select_related(
        'guest', 'employee', 'department', 'registered_by'
    )
    total_visits_today = Visit.objects.filter(entry_time__date=timezone.now().date()).count()
    # Можно добавить другую статистику

    context = {
        'current_visits': current_visits,
        'total_visits_today': total_visits_today,
        # Ссылка на стандартную админку Django для управления пользователями, департаментами и т.д.
        'admin_url': '/admin/'
    }
    # Можно использовать отдельный шаблон или переиспользовать/расширить employee_dashboard
    return render(request, 'visitors/admin_dashboard.html', context)
# ---------------------------------

# --- Представление для статистики ---
# Кэшируем результат на 15 минут (900 секунд)
@cache_page(60 * 15) # <--- Добавляем декоратор кэширования
@login_required
def visit_statistics_view(request):
    # Проверка прав доступа
    if not (request.user.is_staff or request.user.has_perm('visitors.can_view_visit_statistics')):
        raise PermissionDenied

    today = timezone.now().date()
    start_date_30_days = today - timedelta(days=30)

    logger.debug("Attempting to fetch statistics data...")

    # 1. Посещения за последние 30 дней (ОБЪЕДИНЕННЫЕ)
    visits_daily_official = Visit.objects.filter(
        entry_time__date__gte=start_date_30_days
    ).annotate(date=TruncDate('entry_time')).values('date').annotate(count=Count('id')).order_by('date')

    visits_daily_student = StudentVisit.objects.filter(
        entry_time__date__gte=start_date_30_days
    ).annotate(date=TruncDate('entry_time')).values('date').annotate(count=Count('id')).order_by('date')

    # Объединяем в Python
    daily_counts = defaultdict(int)
    # Предзаполняем нули для всех дней в диапазоне
    current_date = start_date_30_days
    while current_date <= today:
        daily_counts[current_date] = 0
        current_date += timedelta(days=1)

    # Добавляем реальные данные
    for v in visits_daily_official:
        if v['date'] in daily_counts: daily_counts[v['date']] += v['count']
    for v in visits_daily_student:
        if v['date'] in daily_counts: daily_counts[v['date']] += v['count']

    # Сортируем по дате и формируем списки для графика
    sorted_daily_counts = sorted(daily_counts.items())
    visits_daily_labels_final = [d.strftime('%Y-%m-%d') for d, count in sorted_daily_counts]
    visits_daily_data_final = [count for d, count in sorted_daily_counts]


    # 2. Посещения по департаментам (ОБЪЕДИНЕННЫЕ)
    # Используем values_list для эффективности
    visits_dept_official = Visit.objects.filter(department__isnull=False).values_list(
        'department__name').annotate(count=Count('id'))
    visits_dept_student = StudentVisit.objects.values_list( # У StudentVisit департамент обязателен
        'department__name').annotate(count=Count('id'))

    dept_counts = defaultdict(int)
    for name, count in visits_dept_official: dept_counts[name] += count
    for name, count in visits_dept_student: dept_counts[name] += count

    # Сортируем и берем топ 10
    sorted_dept_counts = sorted(dept_counts.items(), key=lambda item: item[1], reverse=True)[:10]
    visits_dept_labels = [item[0] for item in sorted_dept_counts]
    visits_dept_data = [item[1] for item in sorted_dept_counts]


    # 3. Посещения по сотрудникам (только для Visit) - остается без изменений
    visits_by_employee = Visit.objects.values(
        'employee__first_name', 'employee__last_name', 'employee__username' # Убрали email, т.к. не используется в метке
    ).annotate(count=Count('id')).order_by('-count')[:10]

    visits_employee_labels = []
    for ve in visits_by_employee:
        # Используем get_full_name() для корректного отображения
        try:
            # Попробуем получить пользователя для get_full_name (может быть неэффективно)
            # Лучше передавать ID и получать имя в шаблоне или JS, если нужно больше деталей
            # Здесь оставляем простой вариант для примера
            user = User.objects.get(username=ve['employee__username']) # Не очень эффективно
            label = user.get_full_name() or user.username
        except User.DoesNotExist:
            label = ve['employee__username'] # Фоллбэк на username
        visits_employee_labels.append(label)
    visits_employee_data = [ve['count'] for ve in visits_by_employee]


    # 4. Средняя продолжительность визита (ОБЪЕДИНЕННАЯ) - Оптимизированный расчет
    # Используем aggregate для расчета среднего в базе данных
    avg_duration_official = Visit.objects.filter(exit_time__isnull=False).aggregate(
        avg_dur=Avg(F('exit_time') - F('entry_time'))
    )['avg_dur'] or timedelta(0)

    avg_duration_student = StudentVisit.objects.filter(exit_time__isnull=False).aggregate(
        avg_dur=Avg(F('exit_time') - F('entry_time'))
    )['avg_dur'] or timedelta(0)

    # Взвешенное среднее (если нужно точное) или простое среднее (если распределение похоже)
    # Простой вариант: среднее арифметическое средних длительностей
    # Более точный: (total_duration_official + total_duration_student) / (count_official + count_student)
    # Оставим простой вариант для примера:
    total_avg_seconds = (avg_duration_official.total_seconds() + avg_duration_student.total_seconds()) / 2 if (avg_duration_official or avg_duration_student) else 0
    avg_duration_minutes = round(total_avg_seconds / 60) if total_avg_seconds > 0 else 0


    # 5. Общее количество визитов
    total_visits_count = Visit.objects.count() + StudentVisit.objects.count()

    context = {
        'visits_daily_labels': json.dumps(visits_daily_labels_final),
        'visits_daily_data': json.dumps(visits_daily_data_final),
        'visits_dept_labels': json.dumps(visits_dept_labels),
        'visits_dept_data': json.dumps(visits_dept_data),
        'visits_employee_labels': json.dumps(visits_employee_labels),
        'visits_employee_data': json.dumps(visits_employee_data),
        'average_duration_minutes': avg_duration_minutes,
        'total_visits_count': total_visits_count,
    }
    logger.info(f"Statistics data prepared for user '{request.user.username}'.")
    return render(request, 'visitors/visit_statistics.html', context)
# ---------------------------------

# --- Представление для регистрации визита студента ---
@login_required
@permission_required('visitors.can_register_student_visit', raise_exception=True)
# `raise_exception=True` вызовет ошибку 403 Forbidden, если пользователь аутентифицирован, но не имеет права.
# Если False (по умолчанию) или не указано, пользователя перенаправит на страницу входа (LOGIN_URL).
# Вы можете также указать login_url: @permission_required('visitors.can_register_student_visit', login_url='/accounts/login/')
def register_student_visit_view(request):
    if request.method == 'POST':
        form = StudentVisitRegistrationForm(request.POST, prefix="student") # Используем префикс
        if form.is_valid():
            try:
                student_visit = form.save(request.user) # Вызываем save без типа регистрации
                try:
                    send_new_visit_notification_to_security(student_visit, 'student')
                except Exception as e:
                    logger.error(f"Ошибка при вызове send_new_visit_notification_to_security для студента: {e}")
                # --------------------------------------------
                messages.success(request, f"Визит студента {student_visit.guest.full_name} успешно зарегистрирован!")
                return redirect('employee_dashboard')
            except Exception as e:
                messages.error(request, f"Ошибка при регистрации визита студента: {e}")
        else:
            messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    else:
        form = StudentVisitRegistrationForm(prefix="student")

    context = {'student_form': form} # Передаем форму под именем student_form
    return render(request, 'visitors/register_student_visit.html', context)
# -------------------------------------------------------------

# --- Представление для экспорта в Excel ---
# --- Представление для экспорта в Excel ---
@login_required
def export_visits_xlsx(request):
    user = request.user # Получаем текущего пользователя
    # --- Проверка прав доступа ---
    #if not (request.user.is_staff or request.user.has_perm('visitors.can_view_visit_statistics')):
        # Можно вернуть ошибку 403 или редирект
        #return HttpResponse("У вас нет прав для экспорта данных.", status=403)
    # ---------------------------
    
    # Эта проверка остается, так как она для общей возможности просмотра статистики/экспорта
    if not (user.is_staff or user.has_perm('visitors.can_view_visit_statistics')):
        # Можно вернуть ошибку 403 или редирект
        return HttpResponse("У вас нет прав для экспорта данных.", status=403)
    # ---------------------------
    
    # Получаем визиты с учетом прав доступа пользователя
    official_visits_qs, student_visits_qs = get_scoped_visits_qs(user)

    # --- Применяем фильтры (логика копируется из visit_history_view) ---
    filter_form = HistoryFilterForm(request.GET or None) # Получаем GET параметры

    # official_visits_qs = Visit.objects.select_related(
    #    'guest', 'employee', 'department', 'registered_by'
    #).all()
    #student_visits_qs = StudentVisit.objects.select_related(
    #    'guest', 'department', 'registered_by'
    #).all()

    # Применяем фильтры, если они есть в запросе
    if request.GET and filter_form.is_valid():
        # Копируем логику фильтрации из visit_history_view
        guest_name = filter_form.cleaned_data.get('guest_name')
        guest_iin = filter_form.cleaned_data.get('guest_iin')
        entry_date_from = filter_form.cleaned_data.get('entry_date_from')
        entry_date_to = filter_form.cleaned_data.get('entry_date_to')
        purpose_filter = filter_form.cleaned_data.get('purpose')
        status = filter_form.cleaned_data.get('status')
        department = filter_form.cleaned_data.get('department')
        employee_info = filter_form.cleaned_data.get('employee_info')
        student_id = filter_form.cleaned_data.get('student_id_number')
        student_group = filter_form.cleaned_data.get('student_group')

        if guest_name:
            official_visits_qs = official_visits_qs.filter(guest__full_name__icontains=guest_name)
            student_visits_qs = student_visits_qs.filter(guest__full_name__icontains=guest_name)
        if guest_iin:
            import hashlib
            iin_hash = hashlib.sha256(guest_iin.encode()).hexdigest()
            official_visits_qs = official_visits_qs.filter(guest__iin_hash=iin_hash)
            student_visits_qs = student_visits_qs.filter(guest__iin_hash=iin_hash)
        if entry_date_from:
            official_visits_qs = official_visits_qs.filter(
                Q(entry_time__date__gte=entry_date_from) |
                Q(entry_time__isnull=True, expected_entry_time__date__gte=entry_date_from)
            )
            student_visits_qs = student_visits_qs.filter(entry_time__date__gte=entry_date_from)
        if entry_date_to:
            official_visits_qs = official_visits_qs.filter(
                Q(entry_time__date__lte=entry_date_to) |
                Q(entry_time__isnull=True, expected_entry_time__date__lte=entry_date_to)
            )
            student_visits_qs = student_visits_qs.filter(entry_time__date__lte=entry_date_to)
        if purpose_filter:
            official_visits_qs = official_visits_qs.filter(purpose__icontains=purpose_filter)
            student_visits_qs = student_visits_qs.filter(purpose__icontains=purpose_filter)
        if status:
            official_visits_qs = official_visits_qs.filter(status=status)
            student_visits_qs = student_visits_qs.filter(status=status)
        if department:
            official_visits_qs = official_visits_qs.filter(department=department)
            student_visits_qs = student_visits_qs.filter(department=department)
        if employee_info:
            official_visits_qs = official_visits_qs.filter(
                Q(employee__first_name__icontains=employee_info) |
                Q(employee__last_name__icontains=employee_info) |
                Q(employee__email__icontains=employee_info)
            )
        if student_id:
            student_visits_qs = student_visits_qs.filter(student_id_number__icontains=student_id)
        if student_group:
            student_visits_qs = student_visits_qs.filter(student_group__icontains=student_group)


    # --- Объединение результатов ---
    # Добавляем тип для дальнейшего использования
    for v in official_visits_qs: v.visit_kind = 'official'
    for v in student_visits_qs: v.visit_kind = 'student'

    # Используем ключ сортировки из visit_history_view
    def get_sort_key(visit):
        relevant_time = None
        if visit.entry_time:
            relevant_time = visit.entry_time
        elif hasattr(visit, 'expected_entry_time') and visit.expected_entry_time:
            relevant_time = visit.expected_entry_time
        very_old_time = timezone.make_aware(datetime.datetime.min, timezone.get_default_timezone())
        return relevant_time if relevant_time else very_old_time

    combined_list = sorted(
        chain(official_visits_qs, student_visits_qs),
        key=get_sort_key,
        reverse=True
    )
    # ----------------------------

    # --- Создание Excel файла ---
    wb = Workbook()
    ws = wb.active
    ws.title = "История визитов"

    # Заголовки колонок
    headers = [
        "ID Визита", "Тип визита", "Статус", "ФИО Посетителя", "ИИН", "Телефон Посетителя", "Email Посетителя",
        "Сотрудник ФИО", "Сотрудник Email", "Студент ID", "Студент Группа", "Студент Курс",
        "Департамент", "Цель визита", "Конт. тел. сотрудника (визит)",
        "Планируемое время входа", "Время входа", "Время выхода", "Зарегистрировал ФИО", "Зарегистрировал Email",
        "Согласие ПДн"
    ]
    ws.append(headers)

    # Делаем заголовки жирными (опционально)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Заполняем строки данными
    # Получаем отображения статусов
    status_display_map = dict(Visit._meta.get_field('status').choices)

    for visit in combined_list:
        # Форматируем даты и время
        entry_time_str = visit.entry_time.astimezone(timezone.get_default_timezone()).strftime('%Y-%m-%d %H:%M:%S') if visit.entry_time else ''
        exit_time_str = visit.exit_time.astimezone(timezone.get_default_timezone()).strftime('%Y-%m-%d %H:%M:%S') if visit.exit_time else ''
        expected_entry_time_str = ''
        consent_acknowledged_str = ''

        # Подготавливаем данные в зависимости от типа визита
        row_data = []
        status_str = status_display_map.get(visit.status, visit.status) # Отображение статуса

        if visit.visit_kind == 'official':
            expected_entry_time_str = visit.expected_entry_time.astimezone(timezone.get_default_timezone()).strftime('%Y-%m-%d %H:%M:%S') if visit.expected_entry_time else ''
            consent_acknowledged_str = "Да" if visit.consent_acknowledged else "Нет"
            row_data = [
                visit.id,
                "Гость сотрудника/Другое", # Тип
                status_str, # Статус
                visit.guest.full_name,
                getattr(visit.guest, 'iin', None),
                visit.guest.phone_number,
                visit.guest.email,
                visit.employee.get_full_name() if visit.employee else '-', # Сотрудник ФИО
                visit.employee.email if visit.employee else '-',           # Сотрудник Email
                '-', # Студент ID
                '-', # Студент Группа
                '-', # Студент Курс
                visit.department.name if visit.department else '-',        # Департамент
                visit.purpose,
                visit.employee_contact_phone, # Конт. тел. сотрудника (визит)
                expected_entry_time_str, # Планируемое время
                entry_time_str,
                exit_time_str,
                visit.registered_by.get_full_name() if visit.registered_by else '-',
                visit.registered_by.email if visit.registered_by else '-',
                consent_acknowledged_str, # Согласие
            ]
        elif visit.visit_kind == 'student':
             # У студенческих визитов нет этих полей
             expected_entry_time_str = '-'
             consent_acknowledged_str = '-'
             row_data = [
                visit.id,
                "Студент/Абитуриент", # Тип
                status_str, # Статус
                visit.guest.full_name,
                getattr(visit.guest, 'iin', None),
                visit.guest.phone_number,
                visit.guest.email,
                '-', # Сотрудник ФИО
                '-', # Сотрудник Email
                visit.student_id_number, # Студент ID
                visit.student_group,     # Студент Группа
                visit.student_course,    # Студент Курс
                visit.department.name,   # Департамент (обязателен для StudentVisit)
                visit.purpose,
                 '-', # Конт. тел. сотрудника (визит) - нет в этой модели
                expected_entry_time_str, # Планируемое время
                entry_time_str,
                exit_time_str,
                visit.registered_by.get_full_name() if visit.registered_by else '-',
                visit.registered_by.email if visit.registered_by else '-',
                consent_acknowledged_str, # Согласие
            ]
        else: # На случай непредвиденных данных
            row_data = [visit.id, 'Неизвестный тип'] + ['-'] * (len(headers) - 2)

        # Добавляем строку в лист Excel
        # Заменяем None на пустую строку для корректного отображения
        ws.append([str(item) if item is not None else '' for item in row_data])

    # Опционально: Настройка ширины колонок
    for col_idx, column_title in enumerate(headers, 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        # Простая автоширина (может быть не идеальной)
        # ws.column_dimensions[column_letter].width = len(column_title) + 5
        # Или установим фиксированную ширину для некоторых колонок
        if column_title in ["ФИО Посетителя", "Цель визита", "Сотрудник ФИО", "Зарегистрировал ФИО"]:
            ws.column_dimensions[column_letter].width = 30
        elif column_title in ["Время входа", "Время выхода", "Планируемое время входа", "Email Посетителя", "Сотрудник Email", "Зарегистрировал Email"]:
            ws.column_dimensions[column_letter].width = 20
        elif column_title in ["ИИН", "Телефон Посетителя", "Конт. тел. сотрудника (визит)"]:
            ws.column_dimensions[column_letter].width = 18
        else:
            ws.column_dimensions[column_letter].width = 15


    # --- Создание HTTP ответа с файлом ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Формируем имя файла с датой и временем
    filename = f"visit_history_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Сохраняем книгу Excel в ответ
    wb.save(response)

    return response
# ------------------------------------------

@login_required
def employee_autocomplete_view(request):
    # Простое rate limit: не чаще 5 запросов/10 сек на IP
    try:
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        cache_key = f"autocomplete_rl_{ip}"
        from django.core.cache import cache
        cnt = cache.get(cache_key, 0)
        if cnt and int(cnt) >= 25:
            return JsonResponse({'results': [], 'count': 0}, status=429)
        cache.incr(cache_key) if cache.get(cache_key) else cache.set(cache_key, 1, timeout=10)
    except Exception:
        pass
    print(f"--- Autocomplete Request ---") # Отладочное сообщение
    print(f"GET params: {request.GET}")   # Печатаем все GET параметры

    term = request.GET.get('term', '').strip()
    department_id = request.GET.get('department_id') # Получаем ID
    print(f"Search Term: '{term}', Department ID: '{department_id}'") # Печатаем полученные значения

    results = []
    queryset = User.objects.filter(is_active=True)

    if len(term) >= 1:
        queryset = queryset.filter(
            Q(first_name__icontains=term) |
            Q(last_name__icontains=term) |
            Q(email__icontains=term)
        )

    print(f"Users found before department filter: {queryset.count()}") # Сколько нашли до фильтра по деп.

    if department_id and department_id.isdigit():
        try:
            dept_id_int = int(department_id)
            # Важный фильтр:
            queryset = queryset.filter(employee_profile__department__id=dept_id_int)
            print(f"Users found AFTER department filter (Dept ID: {dept_id_int}): {queryset.count()}") # Сколько осталось ПОСЛЕ фильтра
        except ValueError:
            print(f"Ошибка: department_id '{department_id}' не является числом.")
            queryset = User.objects.none()
    elif not department_id:
        print("Department ID not provided, returning empty.")
        queryset = User.objects.none()
    else: # На случай если department_id не None, не пустой, но и не isdigit()
         print(f"Department ID '{department_id}' is not a digit, returning empty.")
         queryset = User.objects.none()


    queryset = queryset.order_by('last_name', 'first_name')[:20]
    count_total = queryset.count()

    results = [{'id': user.id, 'text': f"{user.get_full_name() or user.username} ({user.email})"} for user in queryset]

    print(f"Returning {len(results)} results.") # Сколько результатов возвращаем
    print(f"----------------------------")

    return JsonResponse({'results': results, 'count': count_total})
# ---------------------------------


# Кэшируем ответ на 2 часа (60 * 60 * 2)
# Важно: cache_page должен быть ВНЕШНИМ декоратором (применяется последним)
@cache_page(3600 * 2)
@login_required
def get_department_details_view(request):
    """Возвращает детали департамента (имя директора) по ID."""
    department_id = request.GET.get('department_id')
    logger.debug(f"Request for department details received for dept_id: {department_id}")
    data = {'director_name': 'Не назначен или не найден'} # По умолчанию

    if department_id and department_id.isdigit():
        try:
            # Запрашиваем департамент и связанного директора
            dept = Department.objects.select_related('director').get(pk=int(department_id))
            if dept.director:
                # Формируем имя директора
                director_name = dept.director.get_full_name() or dept.director.username
                data['director_name'] = director_name
            else:
                data['director_name'] = 'Не назначен'
        except Department.DoesNotExist:
            logger.debug(f"Department ID {department_id} not found for detail request.")
        except Exception as e:            
            logger.exception(f"Error fetching department details for dept {department_id}.")
            data['director_name'] = 'Ошибка загрузки'
    else:
        logger.warning(f"Invalid or missing department_id '{department_id}' for detail request.")
        
    logger.debug(f"Returning department details: {data}")
    return JsonResponse(data)

# ---------------------------------

@login_required
def get_employee_details_view(request, user_id):
    """Возвращает детали сотрудника (пока только телефон) по его User ID. """
    print(f"--- Get Employee Details Request ---")
    print(f"Received user_id: {user_id}")
    data = {'phone_number': ''} # Значение по умолчанию
    try:
        # Ищем пользователя и его профиль
        user = User.objects.select_related('employee_profile').get(pk=user_id, is_active=True)
        # Получаем телефон из профиля, если он есть
        if hasattr(user, 'employee_profile') and user.employee_profile.phone_number:
            data['phone_number'] = user.employee_profile.phone_number
    except User.DoesNotExist:
        pass # Пользователь не найден, вернется пустой телефон
    except EmployeeProfile.DoesNotExist:
         pass # У пользователя нет профиля, вернется пустой телефон
    except AttributeError:
         # На случай если employee_profile не настроен правильно
         pass

    return JsonResponse(data)
# -----------------------------------------------

"""def get_employee_details_view(request, user_id):
    print(f"--- Get Employee Details Request ---")
    print(f"Received user_id: {user_id}")
    data = {'phone_number': ''}
    try:
        # Используем select_related для профиля
        user = User.objects.select_related('employee_profile').get(pk=user_id, is_active=True)
        print(f"Found user: {user.username}")
        # Проверяем наличие профиля и телефона в нем
        if hasattr(user, 'employee_profile') and user.employee_profile and user.employee_profile.phone_number:
            print(f"User has profile. Phone found: {user.employee_profile.phone_number}")
            data['phone_number'] = user.employee_profile.phone_number
        elif hasattr(user, 'employee_profile') and user.employee_profile:
             print(f"User {user.username} has profile, but NO phone number.")
        else:
            print(f"User {user.username} has NO employee profile linked.")
    except User.DoesNotExist:
        print(f"User with ID {user_id} not found.")
    except Exception as e:
         print(f"Error fetching employee details for user {user_id}: {e}")
    print(f"Returning data: {data}")
    print(f"----------------------------------")
    return JsonResponse(data)"""
# -----------------------------------------------

# --- View для отметки Check-in ---

@login_required
@require_POST # Разрешаем только POST запросы для безопасности
def check_in_visit(request, visit_kind, visit_id):
    """Отмечает фактический вход для ожидающего визита."""
    # TODO: Добавить проверку прав доступа (например, только staff/охрана)
    # if not request.user.is_staff: raise PermissionDenied

    model = Visit if visit_kind == 'official' else StudentVisit if visit_kind == 'student' else None
    if not model:
        messages.error(request, "Неверный тип визита для Check-in.")
        return redirect(request.META.get('HTTP_REFERER', 'visit_history'))

    # Ищем визит со статусом AWAITING_ARRIVAL
    visit = get_object_or_404(model, pk=visit_id, status=STATUS_AWAITING_ARRIVAL)

    try:
        visit.status = STATUS_CHECKED_IN
        visit.entry_time = timezone.now() # Устанавливаем фактическое время входа
        visit.save()
        messages.success(request, f"Вход для '{visit.guest.full_name}' успешно зарегистрирован.")
    except Exception as e:
        messages.error(request, f"Ошибка при регистрации входа: {e}")

    return redirect(request.META.get('HTTP_REFERER', 'visit_history')) # Возвращаемся назад
# ------------------------------------

# --- View для отмены визита ---
@login_required
@require_POST
def cancel_visit(request, visit_kind, visit_id):
    """Отменяет запланированный визит (статус AWAITING_ARRIVAL)."""
    # Проверка прав доступа (кто может отменять? Админ? Регистратор? Хост?)
    # Дадим права админам и тому, кто регистрировал
    model = Visit if visit_kind == 'official' else StudentVisit if visit_kind == 'student' else None
    if not model:
        messages.error(request, "Неверный тип визита для отмены.")
        return redirect(request.META.get('HTTP_REFERER', 'visit_history'))

    try:
        # Ищем визит со статусом AWAITING_ARRIVAL
        visit = get_object_or_404(model.objects.select_related('guest', 'registered_by'), pk=visit_id, status=STATUS_AWAITING_ARRIVAL)

        # Проверка прав
        can_cancel = request.user.is_staff or visit.registered_by == request.user
        # Можно добавить хоста для Visit: or (visit_kind == 'official' and visit.employee == request.user)
        if not can_cancel:
            messages.error(request, "У вас нет прав для отмены этого визита.")
            return redirect(request.META.get('HTTP_REFERER', 'visit_history'))

        visit.status = STATUS_CANCELLED # Устанавливаем статус "Отменен"
        # Можно добавить поле 'cancellation_reason' или 'cancelled_by', если нужно
        # visit.cancelled_by = request.user
        visit.save(update_fields=['status']) # Обновляем только статус
        messages.success(request, f"Визит для '{visit.guest.full_name}' успешно отменен.")
        logger.info(f"User '{request.user.username}' cancelled {visit_kind} visit ID {visit_id}.")
    except Http404:
        messages.warning(request, f"Визит с ID {visit_id} не найден или уже не ожидает прибытия.")
    except Exception as e:
         logger.error(f"Ошибка при отмене {visit_kind} визита ID {visit_id}: {e}", exc_info=True)
         messages.error(request, f"Ошибка при отмене визита: {e}")

    return redirect(request.META.get('HTTP_REFERER', 'visit_history'))
# ---------------------------------


# --- View для страницы настройки профиля ---
@login_required
def profile_setup_view(request):
    # Получаем или создаем профиль (сигнал должен был создать, но get_or_create надежнее)
    profile, created = EmployeeProfile.objects.get_or_create(user=request.user)

    # Если профиль УЖЕ ПОЛНЫЙ и это GET запрос, а пользователь как-то сюда попал, перенаправляем
    if request.method == 'GET' and profile.phone_number and profile.department:
         # Не показываем сообщение, просто редирект
         # messages.info(request, "Ваш профиль уже настроен.")
        return redirect('employee_dashboard') # Или куда нужно

    if request.method == 'POST':
        form = ProfileSetupForm(request.POST, instance=profile)

        if form.is_valid():
            form.save()
            messages.success(request, "Ваш профиль успешно настроен!")
            logger.info(f"User '{request.user.username}' completed profile setup.")
            # Проверяем параметр next для безопасного редиректа
            next_url = request.GET.get('next')
            # Используем is_safe_url для проверки
            if next_url and url_has_allowed_host_and_scheme(url=next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            else:
                return redirect('employee_dashboard') # Редирект по умолчанию
        else:
            logger.warning(f"Profile setup form invalid for user '{request.user.username}'. Errors: {form.errors.as_json()}")
    else: # GET запрос
        form = ProfileSetupForm(instance=profile) # Показываем форму с текущими данными профиля

    context = {
        'form': form
    }
    return render(request, 'visitors/profile_setup.html', context)
# --------------------------------------------------

# --- Представление для логики защищённых приглашений ---
@login_required
def create_guest_invitation(request):
    """Генерирует токен и ссылку для приглашения гостя напрямую, без формы"""
    # Создаем приглашение сразу с токеном
    invitation = GuestInvitation.objects.create(
        employee=request.user,
        token=uuid.uuid4()
    )
    
    # Генерируем ссылку для гостя
    link = request.build_absolute_uri(reverse('guest_invitation_fill', args=[str(invitation.token)]))
    
    # Показываем ссылку пользователю
    context = {
        'invitation_link': link,
        'invitation': invitation
    }
    return render(request, 'visitors/guest_invitation_create.html', context)

def guest_invitation_fill(request, token):
    invitation = get_object_or_404(GuestInvitation, token=token)
    if invitation.is_filled:
        return render(request, 'visitors/guest_invitation_already_filled.html')
    if request.method == 'POST':
        form = GuestInvitationFillForm(request.POST, request.FILES, instance=invitation)
        if form.is_valid():
            invitation = form.save(commit=False)
            invitation.is_filled = True
            invitation.save()
            return render(request, 'visitors/guest_invitation_filled_success.html')
    else:
        form = GuestInvitationFillForm(instance=invitation)
    return render(request, 'visitors/guest_invitation_fill.html', {'form': form, 'invitation': invitation})

@login_required
def finalize_guest_invitation(request, pk):
    invitation = get_object_or_404(GuestInvitation, pk=pk, is_filled=True, is_registered=False)
    if request.method == 'POST':        
        form = GuestInvitationFinalizeForm(request.POST, instance=invitation)
        if form.is_valid():
            invitation = form.save(commit=False)
            # Fix for the MultipleObjectsReturned issue - use filter().first() instead of get_or_create
            guest = Guest.objects.filter(iin=invitation.guest_iin).first()
            if not guest:
                # Create a new guest if no matching guest was found
                guest = Guest.objects.create(
                    iin=invitation.guest_iin,
                    full_name=invitation.guest_full_name,
                    email=invitation.guest_email,
                    phone_number=invitation.guest_phone
                )
            else:
                # Update the guest's information if they already exist
                guest.full_name = invitation.guest_full_name
                guest.email = invitation.guest_email
                guest.phone_number = invitation.guest_phone
                guest.save()
            visit = Visit.objects.create(
                guest=guest,
                employee=invitation.employee,
                department=invitation.employee.employee_profile.department,
                purpose='Гостевой визит по приглашению',
                expected_entry_time=invitation.visit_time,
                registered_by=request.user,
                employee_contact_phone=invitation.employee.employee_profile.phone_number,
                consent_acknowledged=True,
            )
            invitation.is_registered = True
            invitation.visit = visit
            invitation.save()
            # Отправляем уведомление гостю о регистрации визита
            if invitation.guest_email:
                subject_guest = f"Ваш визит зарегистрирован"
                message_guest = (
                    f"Здравствуйте, {invitation.guest_full_name}!\n\n"
                    f"Ваш визит к {invitation.employee.get_full_name() or invitation.employee.username} "
                    f"({invitation.employee.email}) зарегистрирован на {invitation.visit_time.strftime('%Y-%m-%d %H:%M')}.\n"
                    f"Если у вас возникнут вопросы, свяжитесь с сотрудником по телефону: {invitation.guest_phone or '-'}.\n"
                    "Спасибо, что воспользовались нашей системой."
                )
                try:
                    send_mail(
                        subject_guest,
                        message_guest,
                        settings.DEFAULT_FROM_EMAIL,
                        [invitation.guest_email],
                        fail_silently=False
                    )
                except Exception as e:
                    logger.error(f"Ошибка при отправке уведомления гостю {invitation.guest_email}: {e}")
                messages.success(request, 'Визит успешно зарегистрирован.')
                return redirect('employee_dashboard')
    else:
        form = GuestInvitationFinalizeForm(instance=invitation)
    return render(request, 'visitors/guest_invitation_finalize.html', {'form': form, 'invitation': invitation})
# --------------------------------------------------

def cached_static_serve(request, path, **kwargs):
    return cache_control(
        max_age=86400,  # 1 day in seconds
        immutable=True, 
        public=True
    )(serve_static)(request, path, **kwargs)
    
def manifest_json_view(request):
    """
    Serve the manifest.json file with proper content type and encoding
    """
    # Path to our custom manifest.json file
    manifest_path = os.path.join(settings.BASE_DIR, 'static', 'manifest.json')
    
    if os.path.exists(manifest_path):
        with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
            manifest_data = json.load(manifest_file)
            return JsonResponse(manifest_data)
    else:
        # Fallback: generate a simple valid manifest
        manifest_data = {
            "name": "AITU Visitor Pass",
            "short_name": "Visitor Pass",
            "description": "Система управления пропусками для Astana IT University",
            "start_url": "/",
            "display": "standalone",
            "background_color": "#ffffff",
            "theme_color": "#206bc4",
            "orientation": "any",
            "scope": "/",
            "icons": [
                {
                    "src": "/static/img/icons/icon-192x192.png",
                    "sizes": "192x192",
                    "type": "image/png",
                    "purpose": "any maskable"
                },
                {
                    "src": "/static/img/icons/icon-512x512.png",
                    "sizes": "512x512",
                    "type": "image/png",
                    "purpose": "any maskable"
                }
            ]
        }
        return JsonResponse(manifest_data)

def service_worker_view(request):
    """
    Serve the service worker file with UTF-8 encoding to avoid encoding issues on Windows.
    """
    # Path to the service worker file
    service_worker_path = settings.PWA_SERVICE_WORKER_PATH
    
    if os.path.exists(service_worker_path):
        with open(service_worker_path, 'r', encoding='utf-8') as serviceworker_file:
            return HttpResponse(
                serviceworker_file.read(),
                content_type="application/javascript",
            )
    else:
        # Return an empty service worker if the file doesn't exist
        return HttpResponse("// Service worker not found", content_type="application/javascript")

@login_required
def create_group_invitation_view(request):
    """
    Генерирует ссылку для группового визита без параметров группы (департамент, цель, время).
    Параметры группы будут заданы позже при регистрации группового визита.
    """
    # Создаем приглашение сразу с токеном без запроса параметров группы
    group_invitation = GroupInvitation.objects.create(
        employee=request.user,
        token=uuid.uuid4()
    )
    
    # Генерируем ссылку для гостей
    link = request.build_absolute_uri(reverse('group_invitation_fill', args=[str(group_invitation.token)]))
    
    # Показываем ссылку пользователю
    context = {
        'invitation_link': link,
        'group_invitation': group_invitation
    }
    return render(request, 'visitors/groups/create_group_invitation.html', context)


def group_invitation_fill_view(request, token):
    """
    Страница для заполнения данных гостей по ссылке (до 10 человек).
    """
    group_invitation = get_object_or_404(GroupInvitation, token=token)
    GroupGuestFormSet = modelformset_factory(
        GroupGuest,
        fields=('full_name', 'email', 'phone_number', 'iin', 'photo'),
        extra=max(0, 10 - group_invitation.guests.count()),
        max_num=10,
        can_delete=False
    )
        
    if request.method == 'POST':        
        formset = GroupGuestFormSet(request.POST, request.FILES, queryset=group_invitation.guests.all())
        # Получаем данные о визите из формы
        visit_time = request.POST.get('visit_time')
        purpose = request.POST.get('purpose')
        # Проверяем наличие обязательных полей
        if not visit_time or not purpose:
            from django.contrib import messages
            messages.error(request, "Необходимо заполнить время прибытия и цель визита")
            formset = GroupGuestFormSet(request.POST, request.FILES, queryset=group_invitation.guests.all())
            return render(request, 'visitors/groups/group_invitation_fill.html', {
                'group_invitation': group_invitation,
                'formset': formset
            })
        
        if formset.is_valid():
            guests = formset.save(commit=False)
            for guest in guests:
                guest.group_invitation = group_invitation
                guest.is_filled = True
                guest.save()
                
            # Сохраняем информацию о визите
            group_invitation.is_filled = True
            group_invitation.visit_time = visit_time
            group_invitation.purpose = purpose
            
            group_invitation.save()
            return render(request, 'visitors/groups/group_invitation_filled_success.html', {'group_invitation': group_invitation})
    else:
        formset = GroupGuestFormSet(queryset=group_invitation.guests.all())
    return render(request, 'visitors/groups/group_invitation_fill.html', {
        'group_invitation': group_invitation,
        'formset': formset
    })

@login_required
def group_visit_card_view(request, pk):
    """
    Карточка группового визита для отображения в дэшборде.
    """
    group_invitation = get_object_or_404(GroupInvitation, pk=pk)
    guests = group_invitation.guests.all()
    return render(request, 'visitors/groups/group_visit_card.html', {
        'group_invitation': group_invitation,
        'guests': guests
    })

# Для отображения групповых визитов в дэшборде сотрудника:
# В employee_dashboard_view добавить:
# group_invitations = GroupInvitation.objects.filter(employee=request.user).order_by('-created_at')
# и передать в контекст

# --------------------------------------------------------------------------

@login_required
def register_group_visit_view(request, invitation_id):
    """
    Регистрирует групповой визит на основе заполненного приглашения.
    Позволяет добавить информацию: название группы, департамент, цель, время визита.
    """
    try:
        invitation = GroupInvitation.objects.get(
            id=invitation_id,
            employee=request.user,
            is_filled=True,
            is_registered=False
        )
    except GroupInvitation.DoesNotExist:
        messages.error(request, "Приглашение не найдено или уже зарегистрировано.")
        return redirect('employee_dashboard')

    # Получаем всех гостей приглашения
    guests = invitation.guests.all()

    if not guests.exists():
        messages.warning(request, "В приглашении нет ни одного гостя. Дождитесь когда гости заполнят свои данные.")
        return redirect('employee_dashboard')

    if request.method == 'POST':
        form = GroupVisitRegistrationForm(request.POST, instance=invitation)
        if form.is_valid():
            try:
                group_invitation = form.save(request.user)
                messages.success(request, "Групповой визит успешно зарегистрирован!")
                return redirect('employee_dashboard')
            except Exception as e:
                messages.error(request, f"Ошибка при регистрации группового визита: {e}")
        else:
            messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    else:
        form = GroupVisitRegistrationForm(instance=invitation)

    context = {
        'form': form,
        'invitation': invitation,
        'guests': guests,
        'title': 'Регистрация группового визита'
    }
    return render(request, 'visitors/groups/register_group_visit.html', context)

# --- View для отметки выхода ГРУППЫ ---
@login_required
@require_POST
def mark_group_exit_view(request, visit_id):
    """Отмечает выход группы (модель GroupInvitation)."""
    try:
        # Пытаемся получить объект GroupInvitation по ID
        group_visit = get_object_or_404(GroupInvitation, pk=visit_id)

        # Проверяем, зарегистрирована ли группа и не отмечен ли уже выход
        if not group_visit.is_registered or group_visit.is_completed:
            messages.warning(request, f"Невозможно отметить выход для группового визита. "
                                      f"Визит либо не зарегистрирован, либо уже завершен.")
            return redirect(request.META.get('HTTP_REFERER', 'visit_history'))

        # Отмечаем выход группы
        group_visit.exit_time = timezone.now()
        group_visit.is_completed = True
        group_visit.save(update_fields=['exit_time', 'is_completed'])

        messages.success(request, f"Выход для группового визита успешно зарегистрирован.")
        return redirect('visit_history')

    except Http404:
        logger.warning(f"Attempted to mark exit for non-existent GroupInvitation ID: {visit_id} by user {request.user.username}")
        messages.error(request, f"Ошибка: Групповой визит с ID {visit_id} не найден. Возможно, он был удален.")
        return redirect('visit_history')

    except Exception as e:
        logger.error(f"Unexpected error marking group exit for visit ID {visit_id}: {e}", exc_info=True)
        messages.error(request, "Произошла непредвиденная ошибка при отметке выхода группы.")
        return redirect(request.META.get('HTTP_REFERER', 'visit_history'))
